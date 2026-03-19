import base64
import json
import os
import random
import re
import time
import tempfile
from datetime import datetime

import cv2
import numpy as np
import openpyxl
from google.genai import types
from PIL import Image
try:
    from rapidocr_onnxruntime import RapidOCR
except ImportError:
    RapidOCR = None

from adb_driver import ADBDriver
from screen_capture import ScreenCapture
from vlm_navigator import VLMNavigator

from dotenv import load_dotenv

load_dotenv()


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RECORDS_DIR = os.path.join(BASE_DIR, "records")
DEFAULT_EXCEL_PATH = os.path.join(BASE_DIR, "price_history.xlsx")
DEFAULT_ADB_PATH = os.path.join(
    os.path.dirname(BASE_DIR),
    "scrcpy-win64-v3.3.4",
    "adb.exe",
)
EXCEL_HEADERS = [
    "获取时间",
    "店铺名称",
    "商品ID",
    "商品名称",
    "商品规格",
    "当前价格",
    "原价",
    "库存",
    "URL编号",
]
os.makedirs(RECORDS_DIR, exist_ok=True)


class TaobaoWorkflowAgent:
    def __init__(self, excel_path=None, status_callback=None, progress_callback=None, cancel_event=None, workflow_steps=None, auto_cleanup=True, coord_groups=None):
        self.excel_path = excel_path or DEFAULT_EXCEL_PATH
        self.status_callback = status_callback
        self.progress_callback = progress_callback
        self.cancel_event = cancel_event
        self.workflow_steps = workflow_steps or []
        self.auto_cleanup = auto_cleanup
        self.coord_groups = coord_groups or {}
        self.used_coord_group_points = {}
        self.used_random_execute_children = {}
        self.last_condition_result = None

        self._log("初始化 VLM 工作流 Agent")
        self._init_excel()

        self.screencap = ScreenCapture("scrcpy")
        self.adb = ADBDriver(adb_path=DEFAULT_ADB_PATH)
        self.navigator = VLMNavigator()
        self.ocr_engine = RapidOCR() if RapidOCR else None
        self.current_product_title = "未知商品"
        self.current_shop_name = "手机淘宝收藏夹"
        self.taobao_package = "com.taobao.taobao"

    def _reset_runtime_state(self):
        self.used_coord_group_points = {}
        self.used_random_execute_children = {}
        self.last_condition_result = None
        self._log("已重置本轮运行的随机记录")

    def _log(self, message):
        print(f"[Agent] {message}")
        if self.status_callback:
            self.status_callback(message)

    def _report_progress(self, index, success, message):
        if self.progress_callback:
            self.progress_callback(index, success, message)

    def _check_cancel(self):
        if self.cancel_event is not None and self.cancel_event.is_set():
            raise RuntimeError("任务已取消")

    def _sleep_with_cancel(self, seconds):
        end_at = time.time() + max(seconds, 0)
        while time.time() < end_at:
            self._check_cancel()
            time.sleep(min(0.2, end_at - time.time()))

    def _init_excel(self):
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Price History"
            ws.append(EXCEL_HEADERS)
            wb.save(self.excel_path)
            self._log(f"已创建 Excel 数据文件: {self.excel_path}")
            return

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        if existing_headers[: len(EXCEL_HEADERS)] != EXCEL_HEADERS:
            self._log("检测到 Excel 表头与预期不一致，补齐标准表头")
            for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            wb.save(self.excel_path)

    def _last_price_for_sku(self, product_name, sku_name):
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in reversed(rows):
            if len(row) < 6:
                continue
            if row[3] == product_name and row[4] == sku_name:
                try:
                    return float(row[5]) if row[5] is not None else None
                except (TypeError, ValueError):
                    return None
        return None

    def append_to_excel(self, product_name, sku_name, current_price, original_price=None, stock=0):
        product_name = product_name or self.current_product_title or "未知商品"
        sku_name = sku_name or "默认规格"
        try:
            current_price = float(current_price)
        except (TypeError, ValueError):
            self._log(f"跳过无法解析价格的 SKU: {sku_name} -> {current_price}")
            return False

        if original_price in ("", None):
            original_price = None
        else:
            try:
                original_price = float(original_price)
            except (TypeError, ValueError):
                original_price = None

        if original_price is not None and original_price <= current_price:
            original_price = None

        try:
            stock = int(stock)
        except (TypeError, ValueError):
            stock = 0

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        ws.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                self.current_shop_name,
                "VLM-FAVORITES",
                product_name,
                sku_name,
                current_price,
                original_price,
                stock,
                "VLM",
            ]
        )
        wb.save(self.excel_path)
        self._log(f"已写入 Excel: {product_name} | {sku_name} | {current_price}")
        return True

    def _normalize_sku_name(self, name):
        text = str(name or "").strip().lower()
        if not text:
            return ""
        replacements = {
            "【": "[",
            "】": "]",
            "（": "(",
            "）": ")",
            "，": ",",
            "：": ":",
            "“": '"',
            "”": '"',
            "＋": "+",
            "﹢": "+",
            "～": "~",
            "—": "-",
            "–": "-",
            "…": "",
            "...": "",
            "▶": "",
            "▼": "",
            "▾": "",
            "∨": "",
            ">": "",
        }
        for src, dst in replacements.items():
            text = text.replace(src, dst)
        text = re.sub(r"\s+", "", text)
        text = re.sub(r"[^0-9a-z\u4e00-\u9fff\[\]\(\)\+\.\-]+", "", text)
        return text

    def _is_reasonable_sku(self, sku):
        name = str(sku.get("name", "")).strip()
        if not name:
            return False

        price = sku.get("price")
        try:
            price = float(price)
        except (TypeError, ValueError):
            return False

        if price <= 0:
            return False

        original_price = sku.get("original_price")
        if original_price in ("", None):
            sku["original_price"] = None
        else:
            try:
                original_price = float(original_price)
            except (TypeError, ValueError):
                original_price = None
            if original_price is not None and original_price <= price:
                original_price = None
            sku["original_price"] = original_price

        sku["name"] = name
        sku["price"] = price
        return True

    def _merge_sku_entry(self, existing, incoming):
        merged = dict(existing)
        if merged.get("price") in (None, "") and incoming.get("price") not in (None, ""):
            merged["price"] = incoming.get("price")
        if merged.get("original_price") in (None, "") and incoming.get("original_price") not in (None, ""):
            merged["original_price"] = incoming.get("original_price")
        if int(merged.get("stock", 0) or 0) <= 0 and int(incoming.get("stock", 0) or 0) > 0:
            merged["stock"] = incoming.get("stock", 0)
        if len(str(incoming.get("name", "")).strip()) > len(str(merged.get("name", "")).strip()):
            merged["name"] = incoming.get("name", merged.get("name"))
        return merged

    def _merge_exact_sku_lists(self, primary, secondary):
        merged = {}
        for source in (primary or [], secondary or []):
            for sku in source:
                if not self._is_reasonable_sku(sku):
                    continue
                normalized_name = self._normalize_sku_name(sku.get("name"))
                if not normalized_name:
                    continue
                try:
                    price_key = f"{float(sku.get('price')):.2f}"
                except (TypeError, ValueError):
                    price_key = ""
                dedupe_key = f"{normalized_name}|{price_key}"
                incoming = {
                    "name": sku.get("name"),
                    "price": sku.get("price"),
                    "original_price": sku.get("original_price"),
                    "stock": sku.get("stock", 0),
                    "raw_text": sku.get("raw_text", ""),
                }
                if dedupe_key in merged:
                    merged[dedupe_key] = self._merge_sku_entry(merged[dedupe_key], incoming)
                else:
                    merged[dedupe_key] = incoming
        return list(merged.values())

    def _strip_price_text(self, text):
        text = str(text or "").strip()
        if not text:
            return ""
        text = re.sub(r"[¥￥]\s*\d+(?:\.\d+)?", "", text)
        text = re.sub(r"\d+(?:\.\d+)?$", "", text).strip()
        return text.strip(" -—_")

    def _ocr_result_to_items(self, image):
        if self.ocr_engine is None or image is None:
            return []

        try:
            ocr_result, _ = self.ocr_engine(image)
        except Exception as exc:
            self._log(f"OCR 识别失败: {exc}")
            return []

        if not ocr_result:
            return []

        items = []
        for entry in ocr_result:
            if not entry or len(entry) < 2:
                continue
            box = entry[0]
            text = str(entry[1]).strip() if len(entry) > 1 else ""
            score = float(entry[2]) if len(entry) > 2 else 1.0
            if not text:
                continue

            xs = [int(point[0]) for point in box]
            ys = [int(point[1]) for point in box]
            items.append({
                "text": text,
                "score": score,
                "x1": min(xs),
                "y1": min(ys),
                "x2": max(xs),
                "y2": max(ys),
                "cx": (min(xs) + max(xs)) / 2.0,
                "cy": (min(ys) + max(ys)) / 2.0,
                "h": max(1, max(ys) - min(ys)),
            })

        items.sort(key=lambda item: (item["cy"], item["x1"]))
        return items

    def _ocr_result_to_lines(self, image):
        items = self._ocr_result_to_items(image)
        if not items:
            return []

        rows = []
        for item in items:
            matched = None
            for row in rows:
                tolerance = max(18, int(max(row["avg_h"], item["h"]) * 0.7))
                if abs(row["cy"] - item["cy"]) <= tolerance:
                    matched = row
                    break
            if matched is None:
                rows.append({"cy": item["cy"], "avg_h": item["h"], "items": [item]})
            else:
                matched["items"].append(item)
                matched["cy"] = sum(x["cy"] for x in matched["items"]) / len(matched["items"])
                matched["avg_h"] = sum(x["h"] for x in matched["items"]) / len(matched["items"])

        line_items = []
        for row in rows:
            row["items"].sort(key=lambda item: item["x1"])
            line_items.append(row["items"])
        return line_items

    def _extract_skus_by_price_anchors(self, image):
        items = self._ocr_result_to_items(image)
        if not items:
            return []

        image_w = image.shape[1]
        price_regex = re.compile(r"[¥￥]\s*(\d+(?:\.\d+)?)")
        pure_price_regex = re.compile(r"^\d+(?:\.\d+)?$")
        invalid_markers = ("选款建议", "下面是", "加入购物车", "已选", "件商品", "共")
        head_markers = ("现货", "预定", "缺货")

        price_items = []
        for item in items:
            text = item["text"].strip()
            match = price_regex.search(text)
            if not match and pure_price_regex.fullmatch(text) and item["cx"] >= image_w * 0.72:
                match = pure_price_regex.fullmatch(text)
            if not match:
                continue
            try:
                price_value = float(match.group(1))
            except (TypeError, ValueError):
                continue
            if price_value <= 0:
                continue
            price_items.append((item, price_value))

        price_items.sort(key=lambda pair: pair[0]["cy"])
        skus = []
        seen = set()

        for price_item, price_value in price_items:
            tolerance = max(18, int(price_item["h"] * 0.9))
            row_items = [
                item for item in items
                if abs(item["cy"] - price_item["cy"]) <= tolerance and item["x1"] <= price_item["x1"]
            ]
            row_items.sort(key=lambda item: item["x1"])
            texts = [item["text"].strip() for item in row_items if item["text"].strip()]
            if not texts:
                continue

            full_text = "".join(texts)
            if any(marker in full_text for marker in invalid_markers):
                continue

            name_parts = []
            stock = 0
            for item in row_items:
                text = item["text"].strip()
                if item is price_item:
                    stripped = self._strip_price_text(text)
                    if stripped:
                        name_parts.append(stripped)
                    continue
                if text == "缺货":
                    stock = 0
                    continue
                name_parts.append(text)

            name = "".join(name_parts).strip(" -—_")
            name = re.sub(r"[▶▼▾∨]+$", "", name).strip()
            if not name:
                continue
            if not any(name.startswith(marker) for marker in head_markers):
                continue

            normalized_name = self._normalize_sku_name(name)
            if not normalized_name or normalized_name in seen:
                continue
            seen.add(normalized_name)
            skus.append({
                "name": name,
                "price": price_value,
                "original_price": None,
                "stock": stock,
                "raw_text": full_text,
            })

        return skus

    def _extract_skus_by_ocr(self, image):
        anchored_skus = self._extract_skus_by_price_anchors(image)

        lines = self._ocr_result_to_lines(image)
        if not lines:
            return anchored_skus

        line_skus = []
        invalid_markers = ("选款建议", "下面是", "已选", "确定", "加入购物车", "共", "件商品")

        for line in lines:
            texts = [item["text"] for item in line if item["text"]]
            if not texts:
                continue
            full_text = " ".join(texts)
            if any(marker in full_text for marker in invalid_markers):
                continue

            price_idx = None
            price_value = None
            for idx in range(len(line) - 1, -1, -1):
                text = line[idx]["text"]
                match = re.search(r"[¥￥]\s*(\d+(?:\.\d+)?)", text)
                if not match:
                    match = re.fullmatch(r"(\d+(?:\.\d+)?)", text)
                if match:
                    price_idx = idx
                    price_value = float(match.group(1))
                    break
            if price_idx is None or price_value is None:
                continue

            stock = 0
            name_parts = []
            for idx, item in enumerate(line):
                text = item["text"].strip()
                if idx == price_idx:
                    stripped = self._strip_price_text(text)
                    if stripped:
                        name_parts.append(stripped)
                    continue
                if text == "缺货":
                    stock = 0
                    continue
                name_parts.append(text)

            name = "".join(name_parts).strip(" -—_")
            name = re.sub(r"[▶▼▾∨]+$", "", name).strip()
            if not name:
                continue
            if not any(name.startswith(token) for token in ("现货", "预定", "缺货")):
                continue

            line_skus.append({
                "name": name,
                "price": price_value,
                "original_price": None,
                "stock": stock,
                "raw_text": full_text,
            })

        return self._merge_exact_sku_lists(anchored_skus, line_skus)

    def _capture_screen(self):
        self._check_cancel()
        self.screencap.activate_window()
        self._sleep_with_cancel(0.8)
        return self.screencap.capture()

    def cleanup_taobao_session(self):
        """收尾：退出淘宝并回到桌面，确保下次从初始状态启动"""
        try:
            self._log("收尾阶段：退出手机淘宝并清理进程")
            self.adb.stop_app(self.taobao_package)
            self._sleep_with_cancel(1)
            self.adb._run_adb_cmd(["shell", "input", "keyevent", "3"])
            self._sleep_with_cancel(1)
        except Exception as exc:
            self._log(f"清理淘宝会话失败: {exc}")

    def _save_temp_image(self, image, filename):
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        ok, encoded = cv2.imencode(".jpg", image)
        if not ok:
            raise RuntimeError(f"无法编码临时图片: {filename}")
        encoded.tofile(temp_path)
        return temp_path

    def _save_record_image(self, image, prefix, index):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"{prefix}_{timestamp}_{index:02d}.png"
        output_path = os.path.join(RECORDS_DIR, filename)
        ok, encoded = cv2.imencode(".png", image)
        if not ok:
            raise RuntimeError(f"无法保存记录截图: {filename}")
        encoded.tofile(output_path)
        return output_path

    def _parse_point(self, value):
        if value in [None, ""]:
            return None
        if isinstance(value, (list, tuple)) and len(value) == 2:
            return (int(value[0]), int(value[1]))
        if isinstance(value, str) and "," in value:
            x, y = value.split(",", 1)
            return (int(float(x.strip())), int(float(y.strip())))
        return None

    def _random_point_in_rect(self, start_value, end_value):
        p1 = self._parse_point(start_value)
        p2 = self._parse_point(end_value)
        if not p1 or not p2:
            return None
        x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
        x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])
        if x2 <= x1 or y2 <= y1:
            return None
        return (random.randint(x1, x2), random.randint(y1, y2))

    def _parse_coord_group(self, group_name):
        raw_group = self.coord_groups.get(group_name, [])
        if isinstance(raw_group, str):
            raw_items = [item.strip() for item in raw_group.split(";") if item.strip()]
        else:
            raw_items = raw_group

        points = []
        for item in raw_items:
            point = self._parse_point(item)
            if point:
                points.append(point)
        return points

    def _device_point_from_image_point(self, image, point):
        if image is None or not self.adb.screen_size:
            return None
        image_h, image_w = image.shape[:2]
        device_w, device_h = self.adb.screen_size
        scale_x = device_w / image_w
        scale_y = device_h / image_h
        return (int(point[0] * scale_x), int(point[1] * scale_y))

    def _map_editor_point_to_device(self, point):
        image = self._capture_screen()
        if image is None:
            return None
        return self._device_point_from_image_point(image, point)

    def _parse_region_bounds(self, start_value, end_value, image_shape):
        p1 = self._parse_point(start_value)
        p2 = self._parse_point(end_value)
        if not p1 or not p2 or image_shape is None:
            return None
        height, width = image_shape[:2]
        x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
        x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])
        x1 = max(0, min(width, x1))
        x2 = max(0, min(width, x2))
        y1 = max(0, min(height, y1))
        y2 = max(0, min(height, y2))
        if x2 <= x1 or y2 <= y1:
            return None
        return (x1, y1, x2, y2)

    def _crop_image_by_region(self, image, start_value, end_value):
        bounds = self._parse_region_bounds(start_value, end_value, image.shape if image is not None else None)
        if image is None or bounds is None:
            return image, None
        x1, y1, x2, y2 = bounds
        return image[y1:y2, x1:x2], bounds

    def _decode_data_url_image(self, image_data):
        if not image_data:
            return None
        try:
            if "," in image_data:
                _, encoded = image_data.split(",", 1)
            else:
                encoded = image_data
            binary = base64.b64decode(encoded)
            array = np.frombuffer(binary, dtype=np.uint8)
            if array.size == 0:
                return None
            return cv2.imdecode(array, cv2.IMREAD_COLOR)
        except Exception as exc:
            self._log(f"解析判断图片失败: {exc}")
            return None

    def _image_template_match(self, haystack, template):
        if haystack is None or template is None:
            return None
        hay_h, hay_w = haystack.shape[:2]
        tpl_h, tpl_w = template.shape[:2]
        if tpl_h < 4 or tpl_w < 4 or hay_h < tpl_h or hay_w < tpl_w:
            return None

        hay_gray = cv2.cvtColor(haystack, cv2.COLOR_BGR2GRAY)
        tpl_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
        result = cv2.matchTemplate(hay_gray, tpl_gray, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(result)
        return {
            "score": float(max_val),
            "bounds": (max_loc[0], max_loc[1], max_loc[0] + tpl_w, max_loc[1] + tpl_h),
        }

    def _recognize_region_text(self, image):
        if image is None:
            return {"text": "", "items": []}
        items = self._ocr_result_to_items(image)
        texts = [item["text"].strip() for item in items if str(item.get("text", "")).strip()]
        return {"text": " ".join(texts).strip(), "items": items}

    def _compare_condition_value(self, actual_text, expected_text, operator):
        actual_text = str(actual_text or "").strip()
        expected_text = str(expected_text or "").strip()
        numeric_ops = {">", ">=", "<", "<="}
        if operator in numeric_ops:
            try:
                actual_num = float(re.search(r"-?\d+(?:\.\d+)?", actual_text).group(0))
                expected_num = float(expected_text)
            except Exception:
                return False
            if operator == ">":
                return actual_num > expected_num
            if operator == ">=":
                return actual_num >= expected_num
            if operator == "<":
                return actual_num < expected_num
            return actual_num <= expected_num

        if operator == "!=":
            return expected_text not in actual_text if expected_text else bool(actual_text)
        return expected_text in actual_text if expected_text else bool(actual_text)

    def _detect_region_content(self, step):
        image = self._capture_screen()
        if image is None:
            return {"matched": False, "reason": "未获取到 scrcpy 画面"}

        region_image, region_bounds = self._crop_image_by_region(
            image,
            step.get("region_start"),
            step.get("region_end"),
        )
        if region_image is None:
            return {"matched": False, "reason": "截图为空"}

        compare_image = step.get("compare_image") or ""
        compare_value = step.get("compare_value") or ""
        operator = step.get("compare_operator", "=") or "="
        threshold = float(step.get("match_threshold", 0.78) or 0.78)

        if compare_image:
            template = self._decode_data_url_image(compare_image)
            if template is None:
                return {"matched": False, "reason": "判断图片解析失败"}
            match = self._image_template_match(region_image, template)
            if not match:
                return {"matched": False, "reason": "判断图片尺寸不合法"}
            x1, y1, x2, y2 = match["bounds"]
            if region_bounds:
                x1 += region_bounds[0]
                x2 += region_bounds[0]
                y1 += region_bounds[1]
                y2 += region_bounds[1]
            matched = match["score"] >= threshold if operator == "=" else match["score"] < threshold
            return {
                "matched": matched,
                "method": "image",
                "score": round(match["score"], 4),
                "threshold": threshold,
                "raw_text": "",
                "bounds": (x1, y1, x2, y2),
                "tap_start": f"{x1},{y1}",
                "tap_end": f"{x2},{y2}",
                "reason": f"模板匹配得分 {match['score']:.3f}，阈值 {threshold:.2f}",
            }

        recognized = self._recognize_region_text(region_image)
        matched_items = []
        if compare_value:
            expected = str(compare_value).strip()
            for item in recognized["items"]:
                text = str(item.get("text", "")).strip()
                if not text:
                    continue
                if operator == "=" and expected in text:
                    matched_items.append(item)
                elif operator == "!=" and expected not in text:
                    matched_items.append(item)
        bounds = None
        if matched_items:
            x1 = min(item["x1"] for item in matched_items)
            y1 = min(item["y1"] for item in matched_items)
            x2 = max(item["x2"] for item in matched_items)
            y2 = max(item["y2"] for item in matched_items)
            if region_bounds:
                x1 += region_bounds[0]
                x2 += region_bounds[0]
                y1 += region_bounds[1]
                y2 += region_bounds[1]
            bounds = (x1, y1, x2, y2)
        matched = self._compare_condition_value(recognized["text"], compare_value, operator)
        result = {
            "matched": matched,
            "method": "ocr",
            "raw_text": recognized["text"],
            "reason": f"识别文本: {recognized['text'] or '空'}",
        }
        if bounds:
            result["bounds"] = bounds
            result["tap_start"] = f"{bounds[0]},{bounds[1]}"
            result["tap_end"] = f"{bounds[2]},{bounds[3]}"
        return result

    def _click_by_vlm(self, keyword, max_retries=3, wait_after=2.0):
        for attempt in range(1, max_retries + 1):
            self._check_cancel()
            image = self._capture_screen()
            if image is None:
                self._log("未获取到 scrcpy 画面，请确认手机已投屏")
                return False

            point = self.navigator.find_text_center_by_keyword(image, keyword)
            if not point:
                point = self.navigator.find_element_center(image, keyword)
            if point:
                mapped = self._device_point_from_image_point(image, point)
                if mapped is None:
                    return False
                self._log(f"识别到关键字并点击: {keyword}")
                self.adb.tap(mapped[0], mapped[1])
                self._sleep_with_cancel(wait_after)
                return True

            self._log(f"未识别到关键字，重试 {attempt}/{max_retries}: {keyword}")
            self._sleep_with_cancel(1.5)
        return False

    def navigate_to_favorites(self):
        self._log("准备打开淘宝收藏夹")
        if not self.screencap.find_window():
            self._log("未检测到 scrcpy 窗口，进入 Mock 模式")
            return "MOCK"

        self.adb.start_app()
        self._sleep_with_cancel(4)

        foreground_package = self.adb.get_foreground_package()
        if foreground_package != "com.taobao.taobao":
            self._log(f"ADB 启动淘宝未成功前台显示，当前前台: {foreground_package or '未知'}，尝试视觉点击桌面淘宝图标")
            icon_candidates = [
                "桌面或文件夹中橙色背景白色 淘 字样的淘宝图标",
                "生活 文件夹中左侧第一个橙色淘宝图标",
                "手机桌面上淘宝应用图标",
            ]
            opened = False
            for desc in icon_candidates:
                if self._click_by_vlm(desc, max_retries=2, wait_after=4.0):
                    opened = True
                    break
            if not opened:
                self._log("未能通过视觉方式点击淘宝图标")
                return False

        if not self._click_by_vlm("底部导航栏里写着 我的 或 我的淘宝 的按钮", wait_after=2.5):
            return False
        if not self._click_by_vlm("页面中写着 收藏 或 收藏夹 的入口", wait_after=3.0):
            return False

        self._log("已进入淘宝收藏夹")
        return True

    def launch_taobao(self, wait_after=4):
        self._log("启动淘宝")
        self.adb.start_app()
        self._sleep_with_cancel(wait_after)
        return True

    def _extract_product_cards(self, image):
        if not self.navigator.client:
            self._log("缺少 Gemini API Key，无法识别收藏夹商品卡片")
            return []
        temp_path = self._save_temp_image(image, "temp_fav_list.jpg")
        pil_img = Image.open(temp_path)
        prompt = """
返回当前淘宝收藏夹页面里所有可点击商品卡片的中心点。
只返回 JSON 数组，格式必须是 [[x, y], ...]。
不要解释，不要返回 markdown。
"""
        try:
            response = self.navigator.client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[pil_img, prompt],
                config=types.GenerateContentConfig(temperature=0.0),
            )
            match = re.search(r"\[\s*\[.*\]\s*\]", response.text or "", re.DOTALL)
            if not match:
                return []
            points = json.loads(match.group(0))
            return [p for p in points if isinstance(p, list) and len(p) == 2]
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def _extract_title_and_skus_from_image(self, image):
        ocr_skus = self._extract_skus_by_ocr(image)
        if ocr_skus:
            return {"title": self.current_product_title, "shop_name": self.current_shop_name, "skus": ocr_skus}

        if not self.navigator.client:
            self._log("缺少 Gemini API Key，无法识别 SKU 信息")
            return {"title": self.current_product_title, "shop_name": "", "skus": []}
        temp_path = self._save_temp_image(image, "temp_sku_gemini.jpg")
        pil_img = Image.open(temp_path)
        prompt = """
请严格识别这张淘宝商品规格列表截图，返回 JSON 对象。
格式必须是:
{
  "title": "商品标题",
  "shop_name": "店铺名称",
  "skus": [
    {"name": "规格名", "price": 123.45, "original_price": null, "stock": 0}
  ]
}

强约束:
1. 只保留当前截图里真实可见、可以直接读到的 SKU 行，禁止根据上下文猜测或补全。
2. 如果某个 SKU 名称在截图中看不完整、看不清、或你不确定，就不要返回该 SKU。
3. `name` 必须尽量逐字照抄截图中的可见文本，不要改写，不要编造赠品、组合装、顺丰等未看到的内容。
4. `price` 只填写当前这行截图中明确可见的成交价/优惠后价格，只保留数字。
5. `original_price` 只有在当前同一行里明确可见单独的原价/划线价时才填写数字，否则必须返回 null，绝对不要推断。
6. 如果库存看不清，`stock` 填 0。
7. 如果店铺名或标题不确定，可返回空字符串。
8. 只返回 JSON，不要解释。
"""
        try:
            response = self.navigator.client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[pil_img, prompt],
                config=types.GenerateContentConfig(temperature=0.0),
            )
            match = re.search(r"\{.*\}", response.text or "", re.DOTALL)
            if not match:
                return {"title": self.current_product_title, "shop_name": "", "skus": []}
            data = json.loads(match.group(0))
            if not isinstance(data, dict):
                return {"title": self.current_product_title, "shop_name": "", "skus": []}
            data.setdefault("title", self.current_product_title)
            data.setdefault("shop_name", "")
            data.setdefault("skus", [])
            return data
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def _open_sku_panel(self):
        candidates = [
            "商品页中写着 共xx款 或 共X款 的规格展开入口",
            "商品页中写着 选择 规格 数量 的按钮",
            "商品页里用于展开 SKU 列表的按钮",
        ]
        for desc in candidates:
            if self._click_by_vlm(desc, max_retries=2, wait_after=2.5):
                return True
        self._log("未找到 SKU 展开入口")
        return False

    def _scroll_sku_panel(self, image):
        if image is None or not self.adb.screen_size:
            return
        self._check_cancel()
        device_w, device_h = self.adb.screen_size
        self.adb.swipe(
            device_w // 2,
            int(device_h * 0.78),
            device_w // 2,
            int(device_h * 0.32),
        )
        self._sleep_with_cancel(1.3)

    def _resolve_sku_swipe_coords(self, swipe_coords=None, region=None):
        if region:
            p1 = self._parse_point(region.get("start"))
            p2 = self._parse_point(region.get("end"))
            if p1 and p2:
                x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
                x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])
                center_x = int((x1 + x2) / 2)
                return {
                    "x1": center_x,
                    "y1": max(y1 + 40, y2 - 120),
                    "x2": center_x,
                    "y2": min(y2 - 40, y1 + 120),
                }
        return swipe_coords

    def _reset_sku_panel_to_top(self, image, swipe_coords=None, swipe_duration_ms=500, rounds=4):
        if image is None:
            return
        self._log("识别前先尝试将 SKU 列表复位到顶部")
        for _ in range(max(1, int(rounds))):
            self._check_cancel()
            if swipe_coords:
                from_device = self._device_point_from_image_point(
                    image,
                    (int(swipe_coords["x2"]), int(swipe_coords["y2"]))
                )
                to_device = self._device_point_from_image_point(
                    image,
                    (int(swipe_coords["x1"]), int(swipe_coords["y1"]))
                )
                if not from_device or not to_device:
                    break
                self.adb.swipe(
                    from_device[0],
                    from_device[1],
                    to_device[0],
                    to_device[1],
                    int(swipe_duration_ms),
                )
            else:
                device_w, device_h = self.adb.screen_size
                self.adb.swipe(
                    device_w // 2,
                    int(device_h * 0.30),
                    device_w // 2,
                    int(device_h * 0.82),
                    int(swipe_duration_ms),
                )
            self._sleep_with_cancel(1.0)
            image = self._capture_screen()
            if image is None:
                break

    def scan_current_skus(self, open_panel=True, max_scrolls=8, swipe_coords=None, swipe_duration_ms=500, wait_after_scroll=1.3, region=None):
        self._log("开始读取当前商品的 SKU 列表")
        if open_panel:
            self._open_sku_panel()

        effective_swipe_coords = self._resolve_sku_swipe_coords(swipe_coords=swipe_coords, region=region)

        initial_image = self._capture_screen()
        if initial_image is not None:
            self._reset_sku_panel_to_top(initial_image, swipe_coords=effective_swipe_coords, swipe_duration_ms=swipe_duration_ms)

        page_records = []
        for capture_idx in range(max_scrolls):
            image = self._capture_screen()
            if image is None:
                break

            record_path = self._save_record_image(image, "sku_page", capture_idx + 1)
            page_records.append({
                "index": capture_idx + 1,
                "path": record_path,
                "image": image,
            })
            self._log(f"已保存 SKU 截图: {record_path}")

            if effective_swipe_coords:
                from_device = self._device_point_from_image_point(
                    image,
                    (int(effective_swipe_coords["x1"]), int(effective_swipe_coords["y1"]))
                )
                to_device = self._device_point_from_image_point(
                    image,
                    (int(effective_swipe_coords["x2"]), int(effective_swipe_coords["y2"]))
                )
                if not from_device or not to_device:
                    break
                self.adb.swipe(
                    from_device[0],
                    from_device[1],
                    to_device[0],
                    to_device[1],
                    int(swipe_duration_ms),
                )
                self._sleep_with_cancel(wait_after_scroll)
            else:
                self._scroll_sku_panel(image)

        collected = {}
        review_rows = []
        stagnant_rounds = 0

        for page in page_records:
            image = page["image"]
            image_for_extract = image
            if region:
                p1 = self._parse_point(region.get("start"))
                p2 = self._parse_point(region.get("end"))
                if p1 and p2:
                    x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
                    x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])
                    # 给识别范围留出上/下边距，避免把首尾几行 SKU 裁掉
                    x1 = max(0, x1 - 10)
                    y1 = max(0, y1 - 220)
                    x2 = min(image.shape[1], x2 + 10)
                    y2 = min(image.shape[0], y2 + 60)
                    if x2 > x1 and y2 > y1:
                        image_for_extract = image[y1:y2, x1:x2]

            extracted = self._extract_title_and_skus_from_image(image_for_extract)
            if extracted.get("title"):
                self.current_product_title = extracted["title"]
            if extracted.get("shop_name"):
                self.current_shop_name = extracted["shop_name"]

            new_count = 0
            overlap_count = 0
            page_seen = set()
            for sku in extracted.get("skus", []):
                raw_text = str(sku.get("raw_text") or "").strip()
                if not self._is_reasonable_sku(sku):
                    review_rows.append({
                        "page": page["index"],
                        "raw_text": raw_text,
                        "name": str(sku.get("name", "")).strip(),
                        "price": sku.get("price"),
                        "original_price": sku.get("original_price"),
                        "stock": sku.get("stock", 0),
                        "process": "过滤",
                        "process_note": "名称为空或价格无效",
                    })
                    continue
                name = sku["name"]
                normalized_name = self._normalize_sku_name(name)
                if not normalized_name:
                    review_rows.append({
                        "page": page["index"],
                        "raw_text": raw_text,
                        "name": name,
                        "price": sku.get("price"),
                        "original_price": sku.get("original_price"),
                        "stock": sku.get("stock", 0),
                        "process": "过滤",
                        "process_note": "规范化名称为空",
                    })
                    continue
                if normalized_name in page_seen:
                    existing = collected.get(normalized_name, {})
                    review_rows.append({
                        "page": page["index"],
                        "raw_text": raw_text,
                        "name": name,
                        "price": sku.get("price"),
                        "original_price": sku.get("original_price"),
                        "stock": sku.get("stock", 0),
                        "process": "重复删除",
                        "process_note": f"与本轮已识别 SKU 重复: {existing.get('name', name)}",
                    })
                    continue
                page_seen.add(normalized_name)
                incoming = {
                    "name": name,
                    "price": sku.get("price"),
                    "original_price": sku.get("original_price"),
                    "stock": sku.get("stock", 0),
                    "raw_text": raw_text,
                    "first_page": page["index"],
                }
                try:
                    price_key = f"{float(sku.get('price')):.2f}"
                except (TypeError, ValueError):
                    price_key = ""
                dedupe_key = f"{normalized_name}|{price_key}"
                if dedupe_key in collected:
                    overlap_count += 1
                    existing = collected[dedupe_key]
                    collected[dedupe_key] = self._merge_sku_entry(collected[dedupe_key], incoming)
                    review_rows.append({
                        "page": page["index"],
                        "raw_text": raw_text,
                        "name": name,
                        "price": sku.get("price"),
                        "original_price": sku.get("original_price"),
                        "stock": sku.get("stock", 0),
                        "process": "重复删除",
                        "process_note": f"与第 {existing.get('first_page', '?')} 页的 {existing.get('name', name)} 同名同价重复",
                    })
                    continue
                collected[dedupe_key] = incoming
                new_count += 1
                review_rows.append({
                    "page": page["index"],
                    "raw_text": raw_text,
                    "name": name,
                    "price": sku.get("price"),
                    "original_price": sku.get("original_price"),
                    "stock": sku.get("stock", 0),
                    "process": "保留",
                    "process_note": "加入最终结果",
                })

            if new_count == 0 and overlap_count > 0:
                stagnant_rounds += 1
            elif new_count == 0:
                stagnant_rounds += 1
            else:
                stagnant_rounds = 0

            if stagnant_rounds >= 2:
                break

        result_skus = []
        seen_exact_rows = set()
        for sku_data in collected.values():
            exact_name = str(sku_data.get("name", "")).strip()
            try:
                exact_price = f"{float(sku_data.get('price')):.2f}"
            except (TypeError, ValueError):
                exact_price = ""
            exact_key = f"{exact_name}|{exact_price}"
            if exact_key in seen_exact_rows:
                review_rows.append({
                    "page": sku_data.get("first_page", 0),
                    "raw_text": sku_data.get("raw_text", ""),
                    "name": exact_name,
                    "price": sku_data.get("price"),
                    "original_price": sku_data.get("original_price"),
                    "stock": sku_data.get("stock", 0),
                    "process": "重复删除",
                    "process_note": f"最终结果中同名同价重复，已删除: {exact_name} @ {exact_price}",
                })
                continue
            seen_exact_rows.add(exact_key)
            result_skus.append({
                "name": sku_data.get("name"),
                "price": sku_data.get("price"),
                "original_price": sku_data.get("original_price"),
                "stock": sku_data.get("stock", 0),
                "raw_text": sku_data.get("raw_text", ""),
                "process": "保留",
                "process_note": "加入最终结果",
            })
        return {
            "title": self.current_product_title,
            "shop_name": self.current_shop_name,
            "skus": result_skus,
            "review_rows": review_rows,
            "record_paths": [page["path"] for page in page_records],
        }

    def collect_current_skus(self, open_panel=True, max_scrolls=8, swipe_coords=None, swipe_duration_ms=500, wait_after_scroll=1.3, region=None):
        scan_result = self.scan_current_skus(
            open_panel=open_panel,
            max_scrolls=max_scrolls,
            swipe_coords=swipe_coords,
            swipe_duration_ms=swipe_duration_ms,
            wait_after_scroll=wait_after_scroll,
            region=region,
        )

        saved_count = 0
        for sku in scan_result["skus"]:
            if self.append_to_excel(
                scan_result["title"],
                sku["name"],
                sku.get("price"),
                original_price=sku.get("original_price"),
                stock=sku.get("stock", 0),
            ):
                saved_count += 1

        self._log(f"当前商品已记录 {saved_count} 条 SKU")
        return saved_count

    def read_and_record_all_skus(self):
        saved_count = self.collect_current_skus()
        self.adb.back()
        self._sleep_with_cancel(1.5)
        return saved_count

    def _perform_mock_scrape(self, max_items):
        self._log("执行 Mock 采集，用于验证后台链路")
        mock_products = [
            ("【Mock测试】修丽可色修精华", "30ml/瓶", 595.0),
            ("【Mock测试】海蓝之谜面霜", "60ml 精华面霜", 2680.0),
            ("【Mock测试】雅诗兰黛小棕瓶", "50ml 第七代", 640.0),
        ]
        total = min(max_items, len(mock_products))
        self.current_shop_name = "淘宝Mock店"
        for idx, (title, sku, price) in enumerate(mock_products[:total], start=1):
            final_price = price + random.choice([-20, 0, 20, 50])
            self.current_product_title = title
            self.append_to_excel(title, sku, final_price, original_price=price, stock=random.randint(10, 200))
            self._report_progress(idx, True, f"Mock 商品已完成: {title}")
            self._sleep_with_cancel(0.3)
        return total

    def _open_random_favorite(self, wait_after=4):
        image = self._capture_screen()
        if image is None:
            return False

        cards = self._extract_product_cards(image)
        if not cards:
            self._log("未识别到收藏夹商品卡片")
            return False

        point = random.choice(cards)
        device_point = self._device_point_from_image_point(image, point)
        if device_point is None:
            return False

        self._log("打开一个随机收藏商品")
        self.adb.tap(device_point[0], device_point[1])
        self._sleep_with_cancel(wait_after)
        return True

    def _random_click_from_group(self, group_name, wait_after=1):
        points = self._parse_coord_group(group_name)
        if not points:
            self._log(f"坐标组为空或无有效坐标: {group_name}")
            return False

        used = self.used_coord_group_points.setdefault(group_name, set())
        candidates = [point for point in points if f"{point[0]},{point[1]}" not in used]
        if not candidates:
            self._log(f"坐标组 {group_name} 已全部点击过")
            return False

        point = random.choice(candidates)
        used.add(f"{point[0]},{point[1]}")
        device_point = self._map_editor_point_to_device(point)
        if not device_point:
            return False

        self._log(
            f"执行乱序点击[{group_name}]: {point[0]},{point[1]} -> 设备坐标 {device_point[0]},{device_point[1]}"
        )
        self.adb.tap(device_point[0], device_point[1])
        self._sleep_with_cancel(wait_after)
        return True

    def _process_random_favorites(self, max_items=1):
        image = self._capture_screen()
        if image is None:
            return {"success": False, "products_processed": 0, "sku_count": 0}

        cards = self._extract_product_cards(image)
        if not cards:
            self._log("未识别到收藏夹商品卡片")
            return {"success": False, "products_processed": 0, "sku_count": 0}

        random.shuffle(cards)
        target_cards = cards[: max_items if max_items > 0 else len(cards)]
        total_sku_count = 0

        for idx, card_point in enumerate(target_cards, start=1):
            self._check_cancel()
            image = self._capture_screen()
            if image is None:
                break

            device_point = self._device_point_from_image_point(image, card_point)
            if device_point is None:
                continue

            self._log(f"开始处理第 {idx} 个收藏商品")
            self.adb.tap(device_point[0], device_point[1])
            self._sleep_with_cancel(4)

            sku_count = self.read_and_record_all_skus()
            total_sku_count += sku_count
            self._report_progress(idx, True, f"商品处理完成，记录 {sku_count} 条 SKU")

            self.adb.back()
            self._sleep_with_cancel(2)

        self._log(f"本轮 VLM 采集结束，共处理 {len(target_cards)} 个商品，记录 {total_sku_count} 条 SKU")
        return {
            "success": True,
            "products_processed": len(target_cards),
            "sku_count": total_sku_count,
        }

    def _execute_workflow_step(self, step, default_max_items):
        action = step.get("action")
        if not step.get("enabled", True):
            return None

        if action == "launch_taobao":
            return self.launch_taobao(wait_after=float(step.get("wait_after", 4)))
        if action == "navigate_to_favorites":
            return self.navigate_to_favorites()
        if action == "click_vlm":
            return self._click_by_vlm(
                step.get("keyword", step.get("target_desc", "")),
                max_retries=int(step.get("retries", 3)),
                wait_after=float(step.get("wait_after", 2)),
            )
        if action == "click_point":
            tap_start = step.get("tap_start")
            tap_end = step.get("tap_end")
            if (not tap_start or not tap_end) and self.last_condition_result:
                tap_start = self.last_condition_result.get("tap_start", tap_start)
                tap_end = self.last_condition_result.get("tap_end", tap_end)
            point = self._random_point_in_rect(tap_start, tap_end)
            if not point:
                return False
            device_point = self._map_editor_point_to_device(point)
            if not device_point:
                return False
            self._log(f"执行范围随机点击: {point[0]},{point[1]} -> 设备坐标 {device_point[0]},{device_point[1]}")
            self.adb.tap(device_point[0], device_point[1])
            self._sleep_with_cancel(float(step.get("wait_after", 1)))
            return True
        if action == "random_click":
            return self._random_click_from_group(
                step.get("coord_group", ""),
                wait_after=float(step.get("wait_after", 1) or 1),
            )
        if action == "swipe":
            from_point = self._parse_point(step.get("from_coord"))
            to_point = self._parse_point(step.get("to_coord"))
            if not from_point or not to_point:
                return False
            from_device = self._map_editor_point_to_device(from_point)
            to_device = self._map_editor_point_to_device(to_point)
            if not from_device or not to_device:
                return False
            self._log(
                "执行坐标滑动: "
                f"{from_point[0]},{from_point[1]} -> {to_point[0]},{to_point[1]} "
                f"(设备 {from_device[0]},{from_device[1]} -> {to_device[0]},{to_device[1]})"
            )
            self.adb.swipe(
                from_device[0],
                from_device[1],
                to_device[0],
                to_device[1],
                int(step.get("duration_ms", 500)),
            )
            self._sleep_with_cancel(float(step.get("wait_after", 1)))
            return True
        if action == "delay":
            delay_ms = step.get("delay_ms")
            if delay_ms is None:
                delay_ms = float(step.get("seconds", 1)) * 1000
            self._sleep_with_cancel(float(delay_ms) / 1000.0)
            return True
        if action == "open_random_favorite":
            return self._open_random_favorite(wait_after=float(step.get("wait_after", 4) or 4))
        if action == "random_execute":
            children = step.get("children", [])
            if not isinstance(children, list) or not children:
                self._log("随机执行未配置子步骤")
                return False

            step_key = step.get("id") or step.get("name") or "random_execute"
            used = self.used_random_execute_children.setdefault(step_key, set())
            candidates = [(idx, child) for idx, child in enumerate(children) if child.get("enabled", True) and idx not in used]
            if not candidates:
                self._log(f"随机执行步骤已无可用子步骤: {step_key}")
                return False

            idx, chosen = random.choice(candidates)
            used.add(idx)
            self._log(f"随机执行命中子步骤: {chosen.get('name', chosen.get('action', idx))}")
            return self._execute_workflow_step(chosen, default_max_items)
        if action == "region_recognize":
            result = self._detect_region_content(step)
            self.last_condition_result = result if result.get("matched") else None
            self._log(result.get("reason", "范围识别已完成"))
            return result
        if action == "if_condition":
            condition_step = {
                "action": step.get("condition_action", "region_recognize"),
                "region_start": step.get("region_start", ""),
                "region_end": step.get("region_end", ""),
                "compare_operator": step.get("compare_operator", "="),
                "compare_value": step.get("compare_value", ""),
                "compare_image": step.get("compare_image", ""),
                "match_threshold": step.get("match_threshold", 0.78),
            }
            condition_result = self._detect_region_content(condition_step)
            if condition_result.get("matched"):
                self.last_condition_result = condition_result
                self._log(f"IF 判断命中: {condition_result.get('reason', '条件成立')}")
                children = step.get("children", [])
                for child in children:
                    if not child.get("enabled", True):
                        continue
                    child_result = self._execute_workflow_step(child, default_max_items)
                    if child_result is False:
                        return {
                            "matched": True,
                            "executed_children": False,
                            "condition_result": condition_result,
                        }
                return {
                    "matched": True,
                    "executed_children": True,
                    "condition_result": condition_result,
                }
            self.last_condition_result = None
            self._log(f"IF 判断未命中，跳过子步骤: {condition_result.get('reason', '条件不成立')}")
            return {
                "matched": False,
                "executed_children": False,
                "condition_result": condition_result,
            }
        if action == "open_sku_panel":
            return self._open_sku_panel()
        if action == "extract_skus":
            return self.collect_current_skus()
        if action == "extract_skus_loop":
            swipe_coords = None
            from_point = self._parse_point(step.get("from_coord"))
            to_point = self._parse_point(step.get("to_coord"))
            if from_point and to_point:
                swipe_coords = {
                    "x1": from_point[0],
                    "y1": from_point[1],
                    "x2": to_point[0],
                    "y2": to_point[1],
                }
            return self.collect_current_skus(
                open_panel=bool(step.get("open_panel", False)),
                max_scrolls=int(step.get("max_scrolls", 8) or 8),
                swipe_coords=swipe_coords,
                swipe_duration_ms=int(step.get("duration_ms", 500) or 500),
                wait_after_scroll=float(step.get("wait_after", 1.3) or 1.3),
                region={"start": step.get("region_start"), "end": step.get("region_end")},
            )
        if action == "back":
            self.adb.back()
            self._sleep_with_cancel(float(step.get("wait_after", 1)))
            return True
        if action == "home":
            self.adb._run_adb_cmd(["shell", "input", "keyevent", "3"])
            self._sleep_with_cancel(float(step.get("wait_after", 1)))
            return True
        if action == "stop_taobao":
            self.adb.stop_app(self.taobao_package)
            self._sleep_with_cancel(float(step.get("wait_after", 1)))
            return True
        if action == "process_random_favorites":
            max_items = int(step.get("max_items", default_max_items) or default_max_items)
            return self._process_random_favorites(max_items=max_items)
        if action == "sku_recognition":
            return self.collect_current_skus()
        if action == "noop":
            return True

        self._log(f"未知流程动作，已跳过: {action}")
        return None

    def run_configured_workflow(self, max_items=1):
        processed_products = 0
        total_sku_count = 0
        for step in self.workflow_steps:
            self._check_cancel()
            step_name = step.get("name") or step.get("action") or "未命名步骤"
            self._log(f"执行流程步骤: {step_name}")
            result = self._execute_workflow_step(step, max_items)
            if isinstance(result, dict):
                processed_products += int(result.get("products_processed", 0) or 0)
                total_sku_count += int(result.get("sku_count", 0) or 0)
                if result.get("success") is False:
                    return {"success": False, "products_processed": processed_products, "sku_count": total_sku_count}
            elif result is False:
                return {"success": False, "products_processed": processed_products, "sku_count": total_sku_count}
            elif isinstance(result, int):
                total_sku_count += result

        return {"success": True, "products_processed": processed_products, "sku_count": total_sku_count}

    def preview_workflow_step(self, step):
        action = step.get("action")
        if action in ["extract_skus", "extract_skus_loop"]:
            swipe_coords = None
            from_point = self._parse_point(step.get("from_coord"))
            to_point = self._parse_point(step.get("to_coord"))
            if from_point and to_point:
                swipe_coords = {
                    "x1": from_point[0],
                    "y1": from_point[1],
                    "x2": to_point[0],
                    "y2": to_point[1],
                }
            scan_result = self.scan_current_skus(
                open_panel=bool(step.get("open_panel", action == "extract_skus")),
                max_scrolls=int(step.get("max_scrolls", 8) or 8),
                swipe_coords=swipe_coords,
                swipe_duration_ms=int(step.get("duration_ms", 500) or 500),
                wait_after_scroll=float(step.get("wait_after", 1.3) or 1.3),
                region={"start": step.get("region_start"), "end": step.get("region_end")},
            )
            return {
                "success": True,
                "preview_type": "sku_preview",
                "title": scan_result["title"],
                "shop_name": scan_result["shop_name"],
                "skus": scan_result["skus"],
                "review_rows": scan_result.get("review_rows", []),
                "record_paths": scan_result.get("record_paths", []),
            }
        if action == "region_recognize":
            result = self._detect_region_content(step)
            if result.get("matched"):
                self.last_condition_result = result
            return {
                "success": True,
                "preview_type": "condition_preview",
                "result": result,
            }
        if action == "if_condition":
            result = self._execute_workflow_step(step, int(step.get("max_items", 1) or 1))
            return {
                "success": True,
                "preview_type": "condition_preview",
                "result": result,
            }

        result = self._execute_workflow_step(step, int(step.get("max_items", 1) or 1))
        return {
            "success": bool(result is not False),
            "preview_type": "action",
            "result": result,
        }

    def scrape_all_favorites_randomly(self, max_items=5):
        try:
            self._reset_runtime_state()
            if self.workflow_steps:
                result = self.run_configured_workflow(max_items=max_items)
                return result

            nav_status = self.navigate_to_favorites()
            if not nav_status:
                self._log("导航收藏夹失败")
                return {"success": False, "products_processed": 0, "sku_count": 0}

            if nav_status == "MOCK":
                processed = self._perform_mock_scrape(max_items)
                return {"success": True, "products_processed": processed, "sku_count": processed}

            return self._process_random_favorites(max_items=max_items)
        except RuntimeError as exc:
            if "任务已取消" in str(exc):
                self._log("VLM 任务已取消")
                return {"success": False, "canceled": True, "products_processed": 0, "sku_count": 0}
            raise
        finally:
            self._reset_runtime_state()
            if self.auto_cleanup:
                self.cleanup_taobao_session()


if __name__ == "__main__":
    agent = TaobaoWorkflowAgent()
    print(agent.scrape_all_favorites_randomly(max_items=3))
